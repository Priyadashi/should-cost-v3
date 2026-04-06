#!/usr/bin/env python3
"""
Tail Spend Report Generator - 6-tab Excel deliverable

Usage:
    python generate_report.py --scores <scores.csv> --pareto <pareto.json>
        --clusters <clusters.json> --savings <savings.json>
        --cleaned <cleaned.csv> [--output <report.xlsx>]

Generates a professional Excel workbook with 6 tabs:
    1. Executive Summary    Key metrics, recommendations, savings overview
    2. Pareto Analysis      Spend distribution, segment breakdown
    3. Supplier Scores      Full 20-factor scoring with composite scores
    4. Consolidation        Cluster details and recommended actions
    5. Action Plan          Prioritized implementation roadmap
    6. Savings Waterfall    Phase-by-phase savings projection
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
NAVY = "17365D"
WHITE = "FFFFFF"
LIGHT_BLUE = "D9E2F3"
LIGHT_GRAY = "F2F2F2"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
DARK_GREEN = "006100"
DARK_YELLOW = "9C6500"
DARK_RED = "9C0006"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=12)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
DATA_FONT = Font(name="Calibri", size=11)
METRIC_FONT = Font(name="Calibri", bold=True, size=14)
LABEL_FONT = Font(name="Calibri", size=10, color="666666")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=NAVY)

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9")
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Conditional fill for CPS scores
CPS_FILLS = {
    "low": PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid"),
    "med": PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid"),
    "high": PatternFill(start_color=RED, end_color=RED, fill_type="solid"),
}
CPS_FONTS = {
    "low": Font(name="Calibri", size=11, color=DARK_GREEN),
    "med": Font(name="Calibri", size=11, color=DARK_YELLOW),
    "high": Font(name="Calibri", size=11, color=DARK_RED),
}


def _apply_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _apply_subheader_row(ws, row, max_col):
    """Apply subheader styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER


def _apply_cps_formatting(cell, value):
    """Apply conditional formatting based on CPS score."""
    try:
        v = float(value)
    except (ValueError, TypeError):
        return
    if v >= 4.0:
        cell.fill = CPS_FILLS["high"]
        cell.font = CPS_FONTS["high"]
    elif v >= 3.0:
        cell.fill = CPS_FILLS["med"]
        cell.font = CPS_FONTS["med"]
    else:
        cell.fill = CPS_FILLS["low"]
        cell.font = CPS_FONTS["low"]


def _auto_width(ws, min_width=10, max_width=35):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_metric_block(ws, row, col, label, value, fmt=""):
    """Write a large metric with label below."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = METRIC_FONT
    cell.alignment = CENTER
    if fmt:
        cell.number_format = fmt
    label_cell = ws.cell(row=row + 1, column=col, value=label)
    label_cell.font = LABEL_FONT
    label_cell.alignment = CENTER


# ===========================================================================
# Tab builders
# ===========================================================================

def build_executive_summary(wb, pareto, savings, clusters, scores_df):
    """Tab 1: Executive Summary."""
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = NAVY

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "TAIL SPEND ANALYSIS — EXECUTIVE SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated {datetime.now().strftime('%B %d, %Y')}"
    ws["A2"].font = LABEL_FONT

    # Key metrics row
    row = 4
    ws.cell(row=row, column=1, value="KEY METRICS").font = SUBHEADER_FONT
    row = 5

    total_spend = pareto.get("total_spend", 0)
    total_suppliers = pareto.get("supplier_count", 0)
    tail_info = pareto.get("segments", {}).get("Tail", {})
    tail_spend = tail_info.get("total_spend", 0)
    tail_suppliers = tail_info.get("supplier_count", 0)
    est_savings = savings.get("total_estimated_savings", 0)

    metrics = [
        ("Total Spend", total_spend, "$#,##0"),
        ("Total Suppliers", total_suppliers, "#,##0"),
        ("Tail Spend", tail_spend, "$#,##0"),
        ("Tail Suppliers", tail_suppliers, "#,##0"),
        ("Tail %", tail_spend / total_spend * 100 if total_spend else 0, "0.0\"%\""),
        ("Est. Savings", est_savings, "$#,##0"),
        ("Savings %", est_savings / tail_spend * 100 if tail_spend else 0, "0.0\"%\""),
    ]

    for i, (label, value, fmt) in enumerate(metrics):
        _write_metric_block(ws, row, i + 1, label, value, fmt)

    # Segment breakdown
    row = 8
    ws.cell(row=row, column=1, value="SPEND SEGMENTATION").font = SUBHEADER_FONT
    row = 9
    headers = ["Segment", "Suppliers", "% of Suppliers", "Spend", "% of Spend", "Avg per Supplier"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(headers))

    for seg in ["Head", "Core", "Tail"]:
        row += 1
        info = pareto.get("segments", {}).get(seg, {})
        sc = info.get("supplier_count", 0)
        sp = info.get("total_spend", 0)
        ws.cell(row=row, column=1, value=seg).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=row, column=2, value=sc)
        ws.cell(row=row, column=3, value=sc / total_suppliers * 100 if total_suppliers else 0)
        ws.cell(row=row, column=3).number_format = "0.0\"%\""
        ws.cell(row=row, column=4, value=sp)
        ws.cell(row=row, column=4).number_format = "$#,##0"
        ws.cell(row=row, column=5, value=info.get("spend_pct", 0))
        ws.cell(row=row, column=5).number_format = "0.0\"%\""
        ws.cell(row=row, column=6, value=info.get("avg_spend_per_supplier", 0))
        ws.cell(row=row, column=6).number_format = "$#,##0"

    # Recommendation breakdown
    row += 2
    ws.cell(row=row, column=1, value="RECOMMENDATION SUMMARY").font = SUBHEADER_FONT
    row += 1
    rec_headers = ["Action", "Supplier Count", "% of Total"]
    for c, h in enumerate(rec_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(rec_headers))

    if scores_df is not None and "recommendation" in scores_df.columns:
        for rec, count in scores_df["recommendation"].value_counts().items():
            row += 1
            ws.cell(row=row, column=1, value=rec)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=count / len(scores_df) * 100)
            ws.cell(row=row, column=3).number_format = "0.0\"%\""

    # Top clusters
    row += 2
    ws.cell(row=row, column=1, value="TOP CONSOLIDATION OPPORTUNITIES").font = SUBHEADER_FONT
    row += 1
    cl_headers = ["Category", "Suppliers", "Spend", "Strategy", "Est. Savings"]
    for c, h in enumerate(cl_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(cl_headers))

    for cluster in clusters[:10]:
        row += 1
        ws.cell(row=row, column=1, value=cluster.get("category", ""))
        ws.cell(row=row, column=2, value=cluster.get("supplier_count", 0))
        ws.cell(row=row, column=3, value=cluster.get("total_spend", 0))
        ws.cell(row=row, column=3).number_format = "$#,##0"
        ws.cell(row=row, column=4, value=cluster.get("strategy", ""))
        # Find matching savings estimate
        matched_savings = 0
        for cs in savings.get("cluster_savings", []):
            if cs["cluster_id"] == cluster["cluster_id"]:
                matched_savings = cs["estimated_savings"]
                break
        ws.cell(row=row, column=5, value=matched_savings)
        ws.cell(row=row, column=5).number_format = "$#,##0"

    ws.freeze_panes = "A4"
    _auto_width(ws)


def build_pareto_tab(wb, cleaned_df, pareto):
    """Tab 2: Pareto Analysis."""
    ws = wb.create_sheet("Pareto Analysis")
    ws.sheet_properties.tabColor = "4472C4"

    # Header
    ws.merge_cells("A1:H1")
    ws["A1"] = "PARETO ANALYSIS — SUPPLIER SPEND DISTRIBUTION"
    ws["A1"].font = TITLE_FONT

    # Data table
    row = 3
    headers = ["Rank", "Supplier", "Total Spend", "% of Total", "Cumulative %", "Segment"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(headers))

    sorted_df = cleaned_df.sort_values("total_spend", ascending=False).reset_index(drop=True)
    total = sorted_df["total_spend"].sum()

    for i, (_, sup) in enumerate(sorted_df.iterrows()):
        row += 1
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=sup["supplier"])
        ws.cell(row=row, column=3, value=sup["total_spend"])
        ws.cell(row=row, column=3).number_format = "$#,##0"
        ws.cell(row=row, column=4, value=sup.get("spend_pct", sup["total_spend"] / total * 100))
        ws.cell(row=row, column=4).number_format = "0.00\"%\""
        ws.cell(row=row, column=5, value=sup.get("cumulative_pct", 0))
        ws.cell(row=row, column=5).number_format = "0.00\"%\""
        seg = sup.get("segment", "")
        seg_cell = ws.cell(row=row, column=6, value=seg)
        if seg == "Tail":
            seg_cell.fill = CPS_FILLS["high"]
            seg_cell.font = CPS_FONTS["high"]
        elif seg == "Core":
            seg_cell.fill = CPS_FILLS["med"]
            seg_cell.font = CPS_FONTS["med"]
        elif seg == "Head":
            seg_cell.fill = CPS_FILLS["low"]
            seg_cell.font = CPS_FONTS["low"]

    ws.freeze_panes = "A4"
    _auto_width(ws)


def build_scores_tab(wb, scores_df):
    """Tab 3: Supplier Scores."""
    ws = wb.create_sheet("Supplier Scores")
    ws.sheet_properties.tabColor = "ED7D31"

    ws.merge_cells("A1:V1")
    ws["A1"] = "SUPPLIER SCORING — 20-FACTOR ANALYSIS"
    ws["A1"].font = TITLE_FONT

    # Select display columns
    display_cols = ["supplier"]
    factor_cols = [c for c in scores_df.columns
                   if c not in ("supplier", "CPS", "SPS", "RS", "recommendation")]
    composite_cols = ["CPS", "SPS", "RS", "recommendation"]
    all_cols = display_cols + factor_cols + composite_cols
    available = [c for c in all_cols if c in scores_df.columns]

    # Header row
    row = 3
    for c, col_name in enumerate(available, 1):
        label = col_name.replace("_", " ").title()
        ws.cell(row=row, column=c, value=label)
    _apply_subheader_row(ws, row, len(available))

    # Data rows
    for _, data_row in scores_df.iterrows():
        row += 1
        for c, col_name in enumerate(available, 1):
            val = data_row.get(col_name, "")
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = DATA_FONT
            cell.alignment = CENTER

            # Apply CPS conditional formatting
            if col_name == "CPS":
                _apply_cps_formatting(cell, val)
            # Format scores to 1 decimal
            if col_name in factor_cols and isinstance(val, (int, float)):
                cell.number_format = "0.0"
            if col_name in ("SPS", "RS") and isinstance(val, (int, float)):
                cell.number_format = "0.000"

    ws.freeze_panes = "B4"
    _auto_width(ws, min_width=8, max_width=20)


def build_clusters_tab(wb, clusters, savings):
    """Tab 4: Consolidation Clusters."""
    ws = wb.create_sheet("Consolidation")
    ws.sheet_properties.tabColor = "A5A5A5"

    ws.merge_cells("A1:H1")
    ws["A1"] = "CONSOLIDATION CLUSTERS"
    ws["A1"].font = TITLE_FONT

    row = 3
    headers = ["Cluster", "Category", "Suppliers", "Total Spend",
               "Avg CPS", "Avg Risk", "Strategy", "Est. Savings"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(headers))

    savings_map = {cs["cluster_id"]: cs for cs in savings.get("cluster_savings", [])}

    for cluster in clusters:
        row += 1
        cid = cluster["cluster_id"]
        ws.cell(row=row, column=1, value=cid)
        ws.cell(row=row, column=2, value=cluster.get("category", ""))
        ws.cell(row=row, column=3, value=cluster.get("supplier_count", 0))

        spend_cell = ws.cell(row=row, column=4, value=cluster.get("total_spend", 0))
        spend_cell.number_format = "$#,##0"

        cps_cell = ws.cell(row=row, column=5, value=cluster.get("avg_cps", 0))
        cps_cell.number_format = "0.00"
        _apply_cps_formatting(cps_cell, cluster.get("avg_cps", 0))

        rs_cell = ws.cell(row=row, column=6, value=cluster.get("avg_rs", 0))
        rs_cell.number_format = "0.000"

        ws.cell(row=row, column=7, value=cluster.get("strategy", ""))

        cs = savings_map.get(cid, {})
        sav_cell = ws.cell(row=row, column=8, value=cs.get("estimated_savings", 0))
        sav_cell.number_format = "$#,##0"

    # Supplier detail section
    row += 2
    ws.cell(row=row, column=1, value="CLUSTER DETAILS — SUPPLIER LISTS").font = SUBHEADER_FONT
    row += 1

    for cluster in clusters:
        row += 1
        ws.cell(row=row, column=1,
                value=f"Cluster {cluster['cluster_id']}: {cluster.get('category', '')}").font = Font(
            name="Calibri", bold=True, size=11)
        suppliers = cluster.get("suppliers", [])
        for sup in suppliers:
            row += 1
            ws.cell(row=row, column=2, value=sup)

    ws.freeze_panes = "A4"
    _auto_width(ws)


def build_action_plan(wb, clusters, savings, scores_df):
    """Tab 5: Action Plan."""
    ws = wb.create_sheet("Action Plan")
    ws.sheet_properties.tabColor = "70AD47"

    ws.merge_cells("A1:I1")
    ws["A1"] = "IMPLEMENTATION ACTION PLAN"
    ws["A1"].font = TITLE_FONT

    # Phase definitions
    phases = [
        ("Phase 1: Quick Wins", "0-3 months", 0.25),
        ("Phase 2: Consolidation", "3-6 months", 0.35),
        ("Phase 3: Renegotiation", "6-9 months", 0.25),
        ("Phase 4: Process Improvement", "9-12 months", 0.15),
    ]

    total_savings = savings.get("total_estimated_savings", 0)
    savings_map = {cs["cluster_id"]: cs for cs in savings.get("cluster_savings", [])}

    row = 3
    for phase_name, timeline, pct in phases:
        # Phase header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1, value=f"{phase_name} ({timeline})")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

        row += 1
        headers = ["Priority", "Category", "Action", "Suppliers Affected",
                    "Current Spend", "Target Savings", "Owner", "Status", "Notes"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _apply_subheader_row(ws, row, len(headers))

        # Assign clusters to phases based on strategy
        phase_clusters = []
        if "Quick" in phase_name:
            phase_clusters = [c for c in clusters
                              if c.get("strategy") in ("Supplier Elimination", "Category Consolidation")
                              and c.get("supplier_count", 0) <= 5]
            if not phase_clusters:
                phase_clusters = clusters[:3]
        elif "Consolidation" in phase_name:
            phase_clusters = [c for c in clusters
                              if c.get("strategy") in ("Category Consolidation", "Volume Bundling")]
        elif "Renegotiation" in phase_name:
            phase_clusters = [c for c in clusters
                              if c.get("strategy") == "Contract Renegotiation"]
        else:
            phase_clusters = [c for c in clusters
                              if c.get("strategy") not in
                              ("Supplier Elimination", "Category Consolidation",
                               "Volume Bundling", "Contract Renegotiation")]

        for i, cluster in enumerate(phase_clusters, 1):
            row += 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=cluster.get("category", ""))
            ws.cell(row=row, column=3, value=cluster.get("strategy", ""))
            ws.cell(row=row, column=4, value=cluster.get("supplier_count", 0))

            spend_cell = ws.cell(row=row, column=5, value=cluster.get("total_spend", 0))
            spend_cell.number_format = "$#,##0"

            cs = savings_map.get(cluster["cluster_id"], {})
            sav_cell = ws.cell(row=row, column=6, value=cs.get("estimated_savings", 0))
            sav_cell.number_format = "$#,##0"

            ws.cell(row=row, column=7, value="[Assign]")
            ws.cell(row=row, column=8, value="Not Started")
            ws.cell(row=row, column=9, value="")

        if not phase_clusters:
            row += 1
            ws.cell(row=row, column=1, value="No clusters assigned to this phase")
            ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, color="999999")

        row += 2  # spacing between phases

    ws.freeze_panes = "A3"
    _auto_width(ws)


def build_savings_waterfall(wb, savings, pareto):
    """Tab 6: Savings Waterfall."""
    ws = wb.create_sheet("Savings Waterfall")
    ws.sheet_properties.tabColor = "FFC000"

    ws.merge_cells("A1:G1")
    ws["A1"] = "SAVINGS WATERFALL — PROJECTED REALIZATION"
    ws["A1"].font = TITLE_FONT

    total_savings = savings.get("total_estimated_savings", 0)
    total_addressable = savings.get("total_addressable_spend", 0)
    total_spend = pareto.get("total_spend", 0)

    # Summary metrics
    row = 3
    summary = [
        ("Total Spend", total_spend, "$#,##0"),
        ("Addressable Tail Spend", total_addressable, "$#,##0"),
        ("Total Estimated Savings", total_savings, "$#,##0"),
        ("Savings % of Addressable", savings.get("savings_pct_of_addressable", 0), "0.0\"%\""),
        ("Savings % of Total Spend", total_savings / total_spend * 100 if total_spend else 0, "0.0\"%\""),
        ("Feasibility Factor", savings.get("feasibility_factor", 0.5), "0.0"),
    ]

    for i, (label, value, fmt) in enumerate(summary):
        ws.cell(row=row + i, column=1, value=label).font = SUBHEADER_FONT
        cell = ws.cell(row=row + i, column=2, value=value)
        cell.number_format = fmt
        cell.font = DATA_FONT

    # Waterfall phases
    row = row + len(summary) + 2
    ws.cell(row=row, column=1, value="SAVINGS BY PHASE").font = SUBHEADER_FONT
    row += 1
    phase_headers = ["Phase", "Timeline", "% of Savings", "Cumulative Savings",
                     "Annualized Run-Rate"]
    for c, h in enumerate(phase_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(phase_headers))

    phases = [
        ("Quick Wins", "0-3 months", 0.25),
        ("Wave 1: Consolidation", "3-6 months", 0.35),
        ("Wave 2: Renegotiation", "6-9 months", 0.25),
        ("Process Improvements", "9-12 months", 0.15),
    ]

    cumulative = 0
    for phase_name, timeline, pct in phases:
        row += 1
        phase_savings = total_savings * pct
        cumulative += phase_savings

        ws.cell(row=row, column=1, value=phase_name)
        ws.cell(row=row, column=2, value=timeline)

        pct_cell = ws.cell(row=row, column=3, value=pct * 100)
        pct_cell.number_format = "0\"%\""

        cum_cell = ws.cell(row=row, column=4, value=cumulative)
        cum_cell.number_format = "$#,##0"

        rate_cell = ws.cell(row=row, column=5, value=total_savings)
        rate_cell.number_format = "$#,##0"

    # Total row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=3, value=100).number_format = "0\"%\""
    ws.cell(row=row, column=4, value=total_savings).number_format = "$#,##0"
    ws.cell(row=row, column=5, value=total_savings).number_format = "$#,##0"
    for c in range(1, 6):
        ws.cell(row=row, column=c).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=row, column=c).border = Border(
            top=Side(style="double", color="000000"))

    # Per-category savings detail
    row += 3
    ws.cell(row=row, column=1, value="SAVINGS BY CATEGORY").font = SUBHEADER_FONT
    row += 1
    cat_headers = ["Category", "Addressable Spend", "Savings Rate",
                   "Feasibility", "Estimated Savings"]
    for c, h in enumerate(cat_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_subheader_row(ws, row, len(cat_headers))

    for cs in savings.get("cluster_savings", []):
        row += 1
        ws.cell(row=row, column=1, value=cs.get("category", ""))
        ws.cell(row=row, column=2, value=cs.get("addressable_spend", 0))
        ws.cell(row=row, column=2).number_format = "$#,##0"
        ws.cell(row=row, column=3, value=cs.get("savings_rate", 0) * 100)
        ws.cell(row=row, column=3).number_format = "0.0\"%\""
        ws.cell(row=row, column=4, value=cs.get("feasibility_factor", 0))
        ws.cell(row=row, column=4).number_format = "0.0"
        ws.cell(row=row, column=5, value=cs.get("estimated_savings", 0))
        ws.cell(row=row, column=5).number_format = "$#,##0"

    ws.freeze_panes = "A3"
    _auto_width(ws)


# ===========================================================================
# Main entry point
# ===========================================================================

def generate_report(scores_path: str, pareto_path: str,
                    clusters_path: str, savings_path: str,
                    cleaned_path: str, output_path: str = "Tail_Spend_Report.xlsx"):
    """Generate the 6-tab Excel report."""
    print("Loading analysis data...")

    scores_df = pd.read_csv(scores_path)
    with open(pareto_path) as f:
        pareto = json.load(f)
    with open(clusters_path) as f:
        clusters = json.load(f)
    with open(savings_path) as f:
        savings = json.load(f)
    cleaned_df = pd.read_csv(cleaned_path)

    print("Building Excel workbook...")
    wb = Workbook()

    print("  Tab 1: Executive Summary")
    build_executive_summary(wb, pareto, savings, clusters, scores_df)

    print("  Tab 2: Pareto Analysis")
    build_pareto_tab(wb, cleaned_df, pareto)

    print("  Tab 3: Supplier Scores")
    build_scores_tab(wb, scores_df)

    print("  Tab 4: Consolidation Clusters")
    build_clusters_tab(wb, clusters, savings)

    print("  Tab 5: Action Plan")
    build_action_plan(wb, clusters, savings, scores_df)

    print("  Tab 6: Savings Waterfall")
    build_savings_waterfall(wb, savings, pareto)

    print(f"Saving to {output_path}...")
    wb.save(output_path)
    print(f"Report saved: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Tail Spend Report Generator")
    parser.add_argument("--scores", required=True, help="Path to supplier_scores.csv")
    parser.add_argument("--pareto", required=True, help="Path to pareto_analysis.json")
    parser.add_argument("--clusters", required=True, help="Path to consolidation_clusters.json")
    parser.add_argument("--savings", required=True, help="Path to savings_estimates.json")
    parser.add_argument("--cleaned", required=True, help="Path to cleaned_spend.csv")
    parser.add_argument("--output", default="Tail_Spend_Report.xlsx", help="Output Excel path")
    args = parser.parse_args()

    try:
        generate_report(args.scores, args.pareto, args.clusters,
                        args.savings, args.cleaned, args.output)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
