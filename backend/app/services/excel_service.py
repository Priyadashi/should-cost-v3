"""
Excel upload/download services for forging cost data.
Parser class pattern - commodity-specific parsers.
"""
import io
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ForgingExcelParser:
    """Parse forging-specific Excel uploads into Part + BOM + Routing data."""

    @staticmethod
    def parse(file_bytes: bytes) -> dict:
        """Parse Excel bytes into structured data. Returns dict with part, bom, routing."""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        result = {
            "part": {"name": "", "part_no": "", "commodity_type": "Forging", "annual_volume": 0},
            "bom": [],
            "routing": [],
            "overheads": {},
            "quoted_price": 0,
            "errors": [],
        }

        try:
            # Try to parse header info from first few rows
            for row in ws.iter_rows(min_row=1, max_row=10, max_col=10, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val = cell.value.strip().lower()
                        # Look for part name
                        if "part name" in val or "product" in val:
                            next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if next_cell.value:
                                result["part"]["name"] = str(next_cell.value)
                        elif "part no" in val or "part number" in val:
                            next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if next_cell.value:
                                result["part"]["part_no"] = str(next_cell.value)
                        elif "volume" in val or "quantity" in val:
                            next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if next_cell.value:
                                try:
                                    result["part"]["annual_volume"] = int(float(str(next_cell.value)))
                                except ValueError:
                                    pass
                        elif "quoted" in val or "price" in val:
                            next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                            if next_cell.value:
                                try:
                                    result["quoted_price"] = float(str(next_cell.value))
                                except ValueError:
                                    pass
        except Exception as e:
            result["errors"].append(f"Header parsing error: {str(e)}")

        return result


def generate_cost_sheet_excel(
    summary: dict,
    line_items: list,
    sensitivity: list,
    volume_analysis: list,
    part_name: str,
    scenario_name: str,
) -> bytes:
    """Generate a formatted Excel workbook for a cost sheet."""
    wb = Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    sub_header_font = Font(name="Calibri", bold=True, size=10)
    sub_header_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    money_format = '#,##0.00'
    pct_format = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="D6D3D1"),
        right=Side(style="thin", color="D6D3D1"),
        top=Side(style="thin", color="D6D3D1"),
        bottom=Side(style="thin", color="D6D3D1"),
    )

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 18

    ws1.append([f"Should-Cost Analysis: {part_name}"])
    ws1.merge_cells("A1:B1")
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color="92400E")

    ws1.append([f"Scenario: {scenario_name}"])
    ws1.append([])

    # Summary table
    summary_rows = [
        ("Total Material (Gross)", summary.get("total_material_gross", 0)),
        ("Scrap Credit", summary.get("total_scrap_credit", 0)),
        ("Total Material (Net)", summary.get("total_material_net", 0)),
        ("Total Conversion", summary.get("total_conversion", 0)),
        ("Total Labor", summary.get("total_labor", 0)),
        ("Total Tooling/NRE", summary.get("total_tooling_nre", 0)),
        ("Total Overhead", summary.get("total_overhead", 0)),
        ("Total SGA", summary.get("total_sga", 0)),
        ("Total Profit", summary.get("total_profit", 0)),
        ("Total Logistics", summary.get("total_logistics", 0)),
        ("", ""),
        ("SHOULD COST", summary.get("should_cost", 0)),
        ("Quoted Price", summary.get("current_price", 0)),
        ("Gap", summary.get("gap", 0)),
        ("Gap %", summary.get("gap_pct", 0) / 100 if summary.get("gap_pct") else 0),
        ("Annual Volume", summary.get("annual_volume", 0)),
        ("Annual Opportunity", summary.get("annual_opportunity", 0)),
    ]

    for label, value in summary_rows:
        ws1.append([label, value])
        row = ws1.max_row
        ws1.cell(row=row, column=1).font = sub_header_font
        if isinstance(value, (int, float)) and label and "%" not in label and "Volume" not in label:
            ws1.cell(row=row, column=2).number_format = money_format
        elif "%" in label:
            ws1.cell(row=row, column=2).number_format = pct_format

    # Sheet 2: Line Items
    ws2 = wb.create_sheet("Detail")
    headers = ["Category", "Item", "Value", "Detail", "Source", "Confidence"]
    ws2.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 20
    ws2.column_dimensions["F"].width = 12

    for item in line_items:
        ws2.append([
            item.get("category", ""),
            item.get("item", ""),
            item.get("value", 0),
            item.get("detail", ""),
            item.get("source", ""),
            item.get("confidence", ""),
        ])
        ws2.cell(row=ws2.max_row, column=3).number_format = money_format

    # Sheet 3: Sensitivity
    ws3 = wb.create_sheet("Sensitivity")
    sens_headers = ["Driver", "New Should Cost", "Impact (₹)", "Impact (%)"]
    ws3.append(sens_headers)
    for i, h in enumerate(sens_headers, 1):
        cell = ws3.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 12

    for s in sensitivity:
        ws3.append([
            s.get("driver", ""),
            s.get("new_should_cost", 0),
            s.get("impact", 0),
            s.get("impact_pct", 0) / 100,
        ])
        ws3.cell(row=ws3.max_row, column=2).number_format = money_format
        ws3.cell(row=ws3.max_row, column=3).number_format = money_format
        ws3.cell(row=ws3.max_row, column=4).number_format = pct_format

    # Volume analysis
    ws3.append([])
    ws3.append(["Volume-Price Analysis"])
    vol_headers = ["Annual Volume", "Batch Size", "Cost/Unit", "Delta", "Delta %"]
    ws3.append(vol_headers)
    for v in volume_analysis:
        ws3.append([
            v.get("annual_volume", 0),
            v.get("batch_size", 0),
            v.get("should_cost_per_unit", 0),
            v.get("delta_vs_base", 0),
            v.get("delta_pct", 0) / 100,
        ])

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_template_excel() -> bytes:
    """Generate a blank template for uploading forging cost data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Data"

    # Header section
    ws.append(["Part Name", ""])
    ws.append(["Part Number", ""])
    ws.append(["Annual Volume", ""])
    ws.append(["Quoted Price (per unit)", ""])
    ws.append([])

    # Material section
    ws.append(["MATERIAL DATA"])
    ws.append(["Grade", "Gross Weight (kg)", "Net Weight (kg)", "Rate (INR/kg)", "Scrap Recovery %"])
    ws.append(["EN8 (Medium Carbon Steel)", 2.8, 1.9, 68, 35])  # Example row

    ws.append([])
    ws.append(["PROCESS ROUTING"])
    ws.append(["Step", "Operation", "Machine", "Cycle Time (min)", "Setup Time (min)", "Batch Size", "Operators", "Labor Rate (INR/hr)", "Tooling (INR/pc)"])
    ws.append([1, "Billet Cutting", "Lathe Machine", 3, 15, 100, 1, 350, 0])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
